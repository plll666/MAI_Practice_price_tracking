from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select
from src.database.models import Products, Shops, ProductLinks, PriceHistory, Subscriptions
from src.core.logger import logger


HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _style_header(ws, headers):
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT


async def build_workbook(user_id: int, db: AsyncSession) -> BytesIO:
    wb = Workbook()

    await _add_products_sheet(wb, user_id, db)
    await _add_shops_sheet(wb, user_id, db)
    await _add_links_sheet(wb, user_id, db)
    await _add_price_history_sheet(wb, user_id, db)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


async def _add_products_sheet(wb: Workbook, user_id: int, db: AsyncSession):
    ws = wb.active
    ws.title = "Товары"
    headers = ["ID", "Название", "Бренд", "Изображение", "Целевая цена", "Дата создания"]
    _style_header(ws, headers)

    query = (
        select(Products.id, Products.title, Products.brand, Products.image_url,
               Subscriptions.target_price, Products.created_at)
        .join(Subscriptions, Subscriptions.product_id == Products.id)
        .where(Subscriptions.user_id == user_id)
        .order_by(Products.id)
    )
    result = await db.execute(query)
    rows = result.all()

    for i, row in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=row.id)
        ws.cell(row=i, column=2, value=row.title)
        ws.cell(row=i, column=3, value=row.brand)
        ws.cell(row=i, column=4, value=row.image_url)
        ws.cell(row=i, column=5, value=float(row.target_price) if row.target_price else None)
        ws.cell(row=i, column=6, value=row.created_at)

    _auto_width(ws, headers)


async def _add_shops_sheet(wb: Workbook, user_id: int, db: AsyncSession):
    ws = wb.create_sheet("Магазины")
    headers = ["ID", "Название", "Домен", "Активен"]
    _style_header(ws, headers)

    query = (
        select(Shops.id, Shops.name, Shops.domain, Shops.is_active)
        .join(ProductLinks, ProductLinks.shop_id == Shops.id)
        .join(Products, Products.id == ProductLinks.product_id)
        .join(Subscriptions, Subscriptions.product_id == Products.id)
        .where(Subscriptions.user_id == user_id)
        .distinct()
        .order_by(Shops.id)
    )
    result = await db.execute(query)
    rows = result.all()

    for i, row in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=row.id)
        ws.cell(row=i, column=2, value=row.name)
        ws.cell(row=i, column=3, value=row.domain)
        ws.cell(row=i, column=4, value="Да" if row.is_active else "Нет")

    _auto_width(ws, headers)


async def _add_links_sheet(wb: Workbook, user_id: int, db: AsyncSession):
    ws = wb.create_sheet("Ссылки")
    headers = ["ID", "Товар", "Магазин", "URL", "Доступен"]
    _style_header(ws, headers)

    query = (
        select(ProductLinks.id, Products.title.label("product_title"),
               Shops.name.label("shop_name"), ProductLinks.url, ProductLinks.is_available)
        .join(Products, Products.id == ProductLinks.product_id)
        .join(Shops, Shops.id == ProductLinks.shop_id)
        .join(Subscriptions, Subscriptions.product_id == Products.id)
        .where(Subscriptions.user_id == user_id)
        .order_by(ProductLinks.id)
    )
    result = await db.execute(query)
    rows = result.all()

    for i, row in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=row.id)
        ws.cell(row=i, column=2, value=row.product_title)
        ws.cell(row=i, column=3, value=row.shop_name)
        ws.cell(row=i, column=4, value=row.url)
        ws.cell(row=i, column=5, value="Да" if row.is_available else "Нет")

    _auto_width(ws, headers)


async def _add_price_history_sheet(wb: Workbook, user_id: int, db: AsyncSession):
    ws = wb.create_sheet("История цен")
    headers = ["ID", "Ссылка (URL)", "Цена", "Дата парсинга"]
    _style_header(ws, headers)

    query = (
        select(PriceHistory.id, ProductLinks.url, PriceHistory.price, PriceHistory.parsed_at)
        .join(ProductLinks, ProductLinks.id == PriceHistory.link_id)
        .join(Products, Products.id == ProductLinks.product_id)
        .join(Subscriptions, Subscriptions.product_id == Products.id)
        .where(Subscriptions.user_id == user_id)
        .order_by(PriceHistory.parsed_at.desc())
    )
    result = await db.execute(query)
    rows = result.all()

    for i, row in enumerate(rows, 2):
        ws.cell(row=i, column=1, value=row.id)
        ws.cell(row=i, column=2, value=row.url)
        ws.cell(row=i, column=3, value=float(row.price) if row.price else None)
        ws.cell(row=i, column=4, value=row.parsed_at)

    _auto_width(ws, headers)


def _auto_width(ws, headers):
    for col, h in enumerate(headers, 1):
        ws.column_dimensions[chr(64 + col)].width = max(len(h) + 2, 15)
